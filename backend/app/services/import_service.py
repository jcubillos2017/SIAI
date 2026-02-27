from __future__ import annotations

import csv
import io
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from beanie import PydanticObjectId

from app.models.computer import Computer
from app.models.import_batch import ImportBatch


EXPECTED_COMPUTER_COLUMNS = [
    "Codigo Inventario",
    "Nombre Equipo",
    "Numero de Serie",
    "Marca",
    "Modelo",
    "Memoria",
    "Tipo de Equipo",
    "Procesador",
    "Tarjeta Video",
    "Disco Duro",
    "Tipo de Adquisicion",
]


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = " ".join(s.split())
    return s


EXPECTED_NORM = [_norm_header(c) for c in EXPECTED_COMPUTER_COLUMNS]


def _read_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = [dict(r) for r in reader]
    return headers, rows


def _read_xlsx(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h is not None else "" for h in header_cells]

    rows: List[Dict[str, Any]] = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row_cells):
            continue
        item: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            val = row_cells[i] if i < len(row_cells) else None
            item[h] = "" if val is None else str(val).strip()
        rows.append(item)

    return headers, rows


def _validate_headers(received_headers: List[str]) -> Tuple[bool, List[str]]:
    received_norm = [_norm_header(h) for h in received_headers]
    missing = [EXPECTED_COMPUTER_COLUMNS[i] for i, en in enumerate(EXPECTED_NORM) if en not in received_norm]
    return (len(missing) == 0), missing


def _pick(row: Dict[str, Any], col_name: str) -> str:
    target = _norm_header(col_name)
    for k, v in row.items():
        if _norm_header(k) == target:
            return (v or "").strip()
    return ""


def _validate_row(row: Dict[str, Any], row_number: int) -> List[dict]:
    errors = []
    inv = _pick(row, "Codigo Inventario")
    host = _pick(row, "Nombre Equipo")
    ser = _pick(row, "Numero de Serie")
    if not inv:
        errors.append({"row": row_number, "field": "Codigo Inventario", "message": "Requerido"})
    if not host:
        errors.append({"row": row_number, "field": "Nombre Equipo", "message": "Requerido"})
    if not ser:
        errors.append({"row": row_number, "field": "Numero de Serie", "message": "Requerido"})
    return errors


async def preview_computers(file_bytes: bytes, filename: str) -> dict:
    if filename.lower().endswith(".xlsx"):
        headers, rows = _read_xlsx(file_bytes)
    else:
        headers, rows = _read_csv(file_bytes)

    ok, missing = _validate_headers(headers)
    errors: List[dict] = []
    if not ok:
        for m in missing:
            errors.append({"row": 0, "field": "HEADER", "message": f"Falta columna: {m}"})

    preview_rows: List[Dict[str, Any]] = []
    for idx, r in enumerate(rows[:10], start=2):
        errors.extend(_validate_row(r, idx))
        preview_rows.append({
            "Codigo Inventario": _pick(r, "Codigo Inventario"),
            "Nombre Equipo": _pick(r, "Nombre Equipo"),
            "Numero de Serie": _pick(r, "Numero de Serie"),
            "Marca": _pick(r, "Marca"),
            "Modelo": _pick(r, "Modelo"),
            "Memoria": _pick(r, "Memoria"),
            "Tipo de Equipo": _pick(r, "Tipo de Equipo"),
            "Procesador": _pick(r, "Procesador"),
            "Tarjeta Video": _pick(r, "Tarjeta Video"),
            "Disco Duro": _pick(r, "Disco Duro"),
            "Tipo de Adquisicion": _pick(r, "Tipo de Adquisicion"),
        })

    return {
        "expected_columns": EXPECTED_COMPUTER_COLUMNS,
        "received_columns": headers,
        "valid": ok and len([e for e in errors if e["row"] == 0]) == 0,
        "preview_rows": preview_rows,
        "errors": errors,
    }


async def commit_computers(file_bytes: bytes, filename: str, user_id: PydanticObjectId) -> dict:
    if filename.lower().endswith(".xlsx"):
        headers, rows = _read_xlsx(file_bytes)
    else:
        headers, rows = _read_csv(file_bytes)

    ok, missing = _validate_headers(headers)
    errors: List[dict] = []
    if not ok:
        for m in missing:
            errors.append({"row": 0, "field": "HEADER", "message": f"Falta columna: {m}"})
        return {"total_rows": 0, "created": 0, "updated": 0, "errors": errors, "batch_id": None}

    created = 0
    updated = 0

    batch = ImportBatch(entity="computers", filename=filename, user_id=user_id, dry_run=False)
    await batch.insert()

    for i, r in enumerate(rows, start=2):
        row_errors = _validate_row(r, i)
        if row_errors:
            errors.extend(row_errors)
            continue

        inv = _pick(r, "Codigo Inventario")
        host = _pick(r, "Nombre Equipo")
        ser = _pick(r, "Numero de Serie")

        doc = await Computer.find_one(Computer.inventory_code == inv)
        now = datetime.utcnow()

        if not doc:
            doc = Computer(
                inventory_code=inv,
                hostname=host,
                serial_number=ser,
                brand=_pick(r, "Marca") or None,
                model=_pick(r, "Modelo") or None,
                memory_raw=_pick(r, "Memoria") or None,
                equipment_type=_pick(r, "Tipo de Equipo") or None,
                cpu=_pick(r, "Procesador") or None,
                gpu=_pick(r, "Tarjeta Video") or None,
                storage_raw=_pick(r, "Disco Duro") or None,
                last_imported_at=now,
                created_at=now,
                updated_at=now,
                acquisition_type=_pick(r, "Tipo de Adquisicion") or None,
            )
            await doc.insert()
            created += 1
        else:
            doc.hostname = host
            doc.serial_number = ser
            doc.brand = _pick(r, "Marca") or None
            doc.model = _pick(r, "Modelo") or None
            doc.memory_raw = _pick(r, "Memoria") or None
            doc.equipment_type = _pick(r, "Tipo de Equipo") or None
            doc.cpu = _pick(r, "Procesador") or None
            doc.gpu = _pick(r, "Tarjeta Video") or None
            doc.storage_raw = _pick(r, "Disco Duro") or None
            doc.last_imported_at = now
            doc.updated_at = now
            await doc.save()
            updated += 1
            doc.acquisition_type = _pick(r, "Tipo de Adquisicion") or None

    batch.total_rows = len(rows)
    batch.created_count = created
    batch.updated_count = updated
    batch.error_count = len(errors)
    await batch.save()

    return {"total_rows": len(rows), "created": created, "updated": updated, "errors": errors, "batch_id": str(batch.id)}


def template_computers_csv_bytes() -> bytes:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(EXPECTED_COMPUTER_COLUMNS)
    writer.writerow(["12345","Prchile-12345","123456789","HP","ProBook","16GB","AIO","Ryzen 7","AMD Radeon","1 tera","Compra"])
    return out.getvalue().encode("utf-8")
